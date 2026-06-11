from __future__ import annotations

import pytest
from fastapi.testclient import TestClient

from backend.api_gateway.base import ServiceContainer, create_app
from backend.auth.roles import AccountStatus, PlatformRole, Principal
from backend.persistence import Database
from backend.services.alei_export_service import ALEIExportService
from backend.services.evaluation_session_service import EvaluationSessionService
from backend.services.lesson_registry_service import LessonRegistryService
from backend.services.prediction_probe_service import PredictionProbeService
from backend.services.study_session_survey_service import StudySessionSurveyService
from backend.services.survey_management_service import SurveyManagementService
from backend.services.survey_response_service import SurveyResponseService


# ── Fake auth ────────────────────────────────────────────────────────────────

class _FakeAuthService:
    def authenticate_token(self, token: str) -> Principal:
        if token == "student-token":
            return Principal(
                id="student-1",
                role=PlatformRole.STUDENT,
                status=AccountStatus.ACTIVE,
            )
        if token == "admin-token":
            return Principal(
                id="admin-1",
                role=PlatformRole.ADMIN,
                status=AccountStatus.ACTIVE,
            )
        raise ValueError("Invalid token")


def _student_headers():
    return {"Authorization": "Bearer student-token"}


def _admin_headers():
    return {"Authorization": "Bearer admin-token"}


# ── Shared fixture ────────────────────────────────────────────────────────────

@pytest.fixture()
def client(tmp_path):
    db = Database(f"sqlite:///{tmp_path}/test.db")
    db.create_schema()
    registry = LessonRegistryService(database=db)

    import backend.lesson_registry as _lr
    _lr.set_registry(registry)

    survey_mgmt = SurveyManagementService(database=db)
    survey_mgmt.seed_defaults()
    survey_resp = SurveyResponseService(database=db)

    svc = ServiceContainer(
        lesson_catalog=None,
        user_evaluation=None,
        auth=_FakeAuthService(),
        execution=None,
        workspace=None,
        metrics_export=None,
        evaluation_session=EvaluationSessionService(database=db),
        prediction_probe=PredictionProbeService(database=db),
        study_session_survey=StudySessionSurveyService(database=db),
        alei_export=ALEIExportService(database=db),
        lesson_registry=registry,
        survey_management=survey_mgmt,
        survey_response=survey_resp,
    )
    app = create_app(services=svc)
    yield TestClient(app), survey_mgmt
    _lr.set_registry(None)  # type: ignore[arg-type]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_template_id(client_tuple, trigger: str) -> str:
    client, mgmt = client_tuple
    template = mgmt.get_template_by_trigger(trigger)
    assert template is not None
    return template.id


def _get_item_ids(client_tuple, trigger: str) -> list[str]:
    client, mgmt = client_tuple
    template = mgmt.get_template_by_trigger(trigger)
    assert template is not None
    items = mgmt.get_items_for_template(template.id)
    return [i.id for i in items]


# ── GET /admin/surveys ────────────────────────────────────────────────────────

def test_list_surveys_admin_returns_all(client):
    tc, _ = client
    response = tc.get("/admin/surveys", headers=_admin_headers())
    assert response.status_code == 200
    body = response.json()
    assert len(body) == 3
    triggers = {s["context_trigger"] for s in body}
    assert triggers == {"post_session", "post_video", "post_replay"}


def test_list_surveys_student_returns_403(client):
    tc, _ = client
    response = tc.get("/admin/surveys", headers=_student_headers())
    assert response.status_code == 403


def test_list_surveys_unauthenticated_returns_401(client):
    tc, _ = client
    response = tc.get("/admin/surveys")
    assert response.status_code == 401


# ── GET /admin/surveys/{id} ───────────────────────────────────────────────────

def test_get_survey_returns_full_template_with_items(client):
    tc, mgmt = client
    template_id = _get_template_id(client, "post_video")
    response = tc.get(f"/admin/surveys/{template_id}", headers=_admin_headers())
    assert response.status_code == 200
    body = response.json()
    assert body["context_trigger"] == "post_video"
    assert [item["question_type"] for item in body["items"]] == [
        "likert_5",
        "likert_5",
        "text",
    ]
    assert body["items"][2]["required"] is False


def test_get_survey_nonexistent_returns_404(client):
    tc, _ = client
    response = tc.get("/admin/surveys/nonexistent-id", headers=_admin_headers())
    assert response.status_code == 404


# ── PATCH /admin/surveys/{id} ─────────────────────────────────────────────────

def test_patch_survey_admin_succeeds(client):
    tc, mgmt = client
    template_id = _get_template_id(client, "post_video")
    response = tc.patch(
        f"/admin/surveys/{template_id}",
        json={"name": "Renamed Video Survey", "is_active": True},
        headers=_admin_headers(),
    )
    assert response.status_code == 200
    body = response.json()
    assert body["name"] == "Renamed Video Survey"


def test_patch_survey_student_returns_403(client):
    tc, mgmt = client
    template_id = _get_template_id(client, "post_video")
    response = tc.patch(
        f"/admin/surveys/{template_id}",
        json={"name": "Hack"},
        headers=_student_headers(),
    )
    assert response.status_code == 403


# ── PATCH /admin/surveys/{id}/items/{item_id} ─────────────────────────────────

def test_patch_non_standardized_item_succeeds(client):
    tc, mgmt = client
    template_id = _get_template_id(client, "post_video")
    item_ids = _get_item_ids(client, "post_video")
    response = tc.patch(
        f"/admin/surveys/{template_id}/items/{item_ids[0]}",
        json={"question_text": "Updated question text."},
        headers=_admin_headers(),
    )
    assert response.status_code == 200
    body = response.json()
    assert body["question_text"] == "Updated question text."


def test_patch_standardized_item_returns_403(client):
    tc, mgmt = client
    template = mgmt.get_template_by_trigger("post_session")
    assert template is not None
    items = mgmt.get_items_for_template(template.id)
    standardized = next(i for i in items if i.is_standardized)
    response = tc.patch(
        f"/admin/surveys/{template.id}/items/{standardized.id}",
        json={"question_text": "Tamper with SUS item."},
        headers=_admin_headers(),
    )
    assert response.status_code == 403


# ── POST /admin/surveys/{id}/items ────────────────────────────────────────────

def test_add_item_returns_201(client):
    tc, mgmt = client
    template_id = _get_template_id(client, "post_replay")
    response = tc.post(
        f"/admin/surveys/{template_id}/items",
        json={"question_text": "Extra question?", "question_type": "text", "order_index": 99, "required": False},
        headers=_admin_headers(),
    )
    assert response.status_code == 201
    body = response.json()
    assert body["question_text"] == "Extra question?"
    assert not body["is_standardized"]


# ── DELETE /admin/surveys/{id}/items/{item_id} ────────────────────────────────

def test_delete_non_standardized_item_returns_204(client):
    tc, mgmt = client
    template_id = _get_template_id(client, "post_video")
    item_ids = _get_item_ids(client, "post_video")
    response = tc.delete(
        f"/admin/surveys/{template_id}/items/{item_ids[0]}",
        headers=_admin_headers(),
    )
    assert response.status_code == 204


def test_delete_standardized_item_returns_403(client):
    tc, mgmt = client
    template = mgmt.get_template_by_trigger("post_session")
    assert template is not None
    items = mgmt.get_items_for_template(template.id)
    standardized = next(i for i in items if i.is_standardized)
    response = tc.delete(
        f"/admin/surveys/{template.id}/items/{standardized.id}",
        headers=_admin_headers(),
    )
    assert response.status_code == 403


# ── GET /surveys/active ───────────────────────────────────────────────────────

def test_list_active_surveys_returns_only_active(client):
    tc, mgmt = client
    response = tc.get("/surveys/active", headers=_student_headers())
    assert response.status_code == 200
    body = response.json()
    assert all(s.get("context_trigger") for s in body)


def test_list_active_surveys_excludes_inactive(client):
    tc, mgmt = client
    template = mgmt.get_template_by_trigger("post_video")
    assert template is not None
    mgmt.update_template(template.id, is_active=False)
    response = tc.get("/surveys/active", headers=_student_headers())
    assert response.status_code == 200
    triggers = {s["context_trigger"] for s in response.json()}
    assert "post_video" not in triggers


# ── GET /surveys/active/{context_trigger} ─────────────────────────────────────

def test_get_active_survey_by_trigger_returns_template_with_items(client):
    tc, _ = client
    response = tc.get("/surveys/active/post_video", headers=_student_headers())
    assert response.status_code == 200
    body = response.json()
    assert body["context_trigger"] == "post_video"
    assert [item["question_type"] for item in body["items"]] == [
        "likert_5",
        "likert_5",
        "text",
    ]
    assert body["items"][2]["required"] is False


def test_get_active_survey_by_trigger_not_found_returns_404(client):
    tc, _ = client
    response = tc.get("/surveys/active/nonexistent_trigger", headers=_student_headers())
    assert response.status_code == 404


# ── POST /surveys/{template_id}/respond ──────────────────────────────────────

def test_submit_response_student_returns_201(client):
    tc, mgmt = client
    template_id = _get_template_id(client, "post_video")
    item_ids = _get_item_ids(client, "post_video")
    response = tc.post(
        f"/surveys/{template_id}/respond",
        json={
            "study_session_id": "sess-1",
            "condition": "adaptive",
            "responses": [
                {"item_id": item_ids[0], "likert_value": 4},
                {"item_id": item_ids[1], "likert_value": 5},
            ],
        },
        headers=_student_headers(),
    )
    assert response.status_code == 201
    body = response.json()
    assert "survey_response_id" in body
    assert body["survey_template_id"] == template_id


def test_submit_response_out_of_range_likert_returns_422(client):
    tc, mgmt = client
    template_id = _get_template_id(client, "post_replay")
    item_ids = _get_item_ids(client, "post_replay")
    response = tc.post(
        f"/surveys/{template_id}/respond",
        json={
            "study_session_id": "sess-1",
            "condition": "control",
            "responses": [
                {"item_id": item_ids[0], "likert_value": 10},
                {"item_id": item_ids[1], "likert_value": 3},
            ],
        },
        headers=_student_headers(),
    )
    assert response.status_code == 422


def test_submit_response_unauthenticated_returns_401(client):
    tc, mgmt = client
    template_id = _get_template_id(client, "post_video")
    item_ids = _get_item_ids(client, "post_video")
    response = tc.post(
        f"/surveys/{template_id}/respond",
        json={
            "study_session_id": "sess-1",
            "condition": "adaptive",
            "responses": [
                {"item_id": item_ids[0], "likert_value": 3},
                {"item_id": item_ids[1], "likert_value": 3},
            ],
        },
    )
    assert response.status_code == 401


# ── ALEI export includes Component Surveys sheet ──────────────────────────────

def test_export_alei_includes_component_surveys_sheet(client):
    import openpyxl
    import io

    tc, mgmt = client
    # Submit one micro-survey response
    template_id = _get_template_id(client, "post_video")
    item_ids = _get_item_ids(client, "post_video")
    tc.post(
        f"/surveys/{template_id}/respond",
        json={
            "study_session_id": "sess-1",
            "condition": "adaptive",
            "responses": [
                {"item_id": item_ids[0], "likert_value": 4},
                {"item_id": item_ids[1], "likert_value": 5},
            ],
        },
        headers=_student_headers(),
    )

    response = tc.get("/evaluation/export/alei-components", headers=_admin_headers())
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert "Component Surveys" in wb.sheetnames
    ws = wb["Component Surveys"]
    rows = list(ws.iter_rows(values_only=True))
    # header + 2 item response rows
    assert len(rows) >= 3
    headers = rows[0]
    assert "context_trigger" in headers
    assert "likert_value" in headers
