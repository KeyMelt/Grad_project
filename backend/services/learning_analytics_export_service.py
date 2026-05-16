from __future__ import annotations

from datetime import datetime, timezone
import io
import json
from typing import Any

from openpyxl import Workbook
from sqlmodel import select

from backend.models.concept_mastery_snapshot import ConceptMasterySnapshot
from backend.models.learning_interaction_event import LearningInteractionEvent
from backend.models.study_buddy_intervention_record import StudyBuddyInterventionRecord
from backend.persistence import Database


class LearningAnalyticsExportService:
    """Builds workbook exports for Study Buddy telemetry, mastery, and interventions."""

    def __init__(self, *, database: Database) -> None:
        self._database = database

    def build_export(self) -> tuple[str, bytes]:
        workbook = Workbook()
        mastery_sheet = workbook.active
        mastery_sheet.title = "Mastery Snapshots"
        mastery_sheet.append(
            [
                "student_id",
                "lesson_id",
                "concept_id",
                "mastery_score",
                "support_need_score",
                "confidence_score",
                "staleness_days",
                "direct_evidence_count",
                "correct_checks",
                "incorrect_checks",
                "intervention_count",
                "passive_signal_count",
                "recorded_at_utc",
                "source_version",
            ]
        )
        interventions_sheet = workbook.create_sheet("Interventions")
        interventions_sheet.append(
            [
                "student_id",
                "lesson_id",
                "concept_id",
                "session_id",
                "trigger_type",
                "trigger_score",
                "intervention_type",
                "status",
                "llm_model",
                "prompt_version",
                "reflection_pass_result",
                "created_at_utc",
                "resolved_at_utc",
                "source_version",
            ]
        )
        telemetry_sheet = workbook.create_sheet("Telemetry Counts")
        telemetry_sheet.append(
            [
                "student_id",
                "lesson_id",
                "concept_id",
                "event_type",
                "event_count",
            ]
        )

        with self._database.session() as session:
            mastery_rows = session.exec(
                select(ConceptMasterySnapshot).order_by(
                    ConceptMasterySnapshot.student_id,
                    ConceptMasterySnapshot.concept_id,
                    ConceptMasterySnapshot.recorded_at_utc,
                )
            ).all()
            for snapshot in mastery_rows:
                evidence = self._decode_json(snapshot.evidence_summary_json)
                mastery_sheet.append(
                    [
                        snapshot.student_id,
                        snapshot.lesson_id,
                        snapshot.concept_id,
                        snapshot.mastery_score,
                        snapshot.support_need_score,
                        snapshot.confidence_score,
                        snapshot.staleness_days,
                        evidence.get("direct_evidence_count", 0),
                        evidence.get("correct_checks", 0),
                        evidence.get("incorrect_checks", 0),
                        evidence.get("intervention_count", 0),
                        evidence.get("passive_signal_count", 0),
                        snapshot.recorded_at_utc.isoformat(),
                        snapshot.source_version,
                    ]
                )

            intervention_rows = session.exec(
                select(StudyBuddyInterventionRecord).order_by(
                    StudyBuddyInterventionRecord.student_id,
                    StudyBuddyInterventionRecord.created_at_utc,
                )
            ).all()
            for record in intervention_rows:
                interventions_sheet.append(
                    [
                        record.student_id,
                        record.lesson_id,
                        record.concept_id,
                        record.session_id,
                        record.trigger_type,
                        record.trigger_score,
                        record.intervention_type,
                        record.status,
                        record.llm_model,
                        record.prompt_version,
                        record.reflection_pass_result,
                        record.created_at_utc.isoformat(),
                        (
                            record.resolved_at_utc.isoformat()
                            if record.resolved_at_utc is not None
                            else None
                        ),
                        record.source_version,
                    ]
                )

            counts: dict[tuple[str, str, str | None, str], int] = {}
            events = session.exec(select(LearningInteractionEvent)).all()
            for event in events:
                key = (
                    event.student_id,
                    event.lesson_id,
                    event.concept_id,
                    event.event_type,
                )
                counts[key] = counts.get(key, 0) + 1
            for (student_id, lesson_id, concept_id, event_type), count in sorted(counts.items()):
                telemetry_sheet.append([student_id, lesson_id, concept_id, event_type, count])

        content = io.BytesIO()
        workbook.save(content)
        content.seek(0)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return f"learning_analytics_{timestamp}.xlsx", content.getvalue()

    def _decode_json(self, raw_value: str) -> dict[str, Any]:
        try:
            decoded = json.loads(raw_value or "{}")
        except json.JSONDecodeError:
            return {}
        return decoded if isinstance(decoded, dict) else {}
