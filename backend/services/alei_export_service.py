from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
import io
import json
from typing import Any

from openpyxl import Workbook
from sqlmodel import select

from backend.config import event_types as ET
from backend.models.learning_interaction_event import LearningInteractionEvent
from backend.models.prediction_probe_attempt import PredictionProbeAttempt
from backend.models.study_session import StudySession
from backend.models.study_session_survey import StudySessionSurvey
from backend.models.survey_response import SurveyItemResponse, SurveyResponse
from backend.models.survey_template import SurveyTemplate, SurveyTemplateItem
from backend.persistence import Database


class ALEIExportService:
    """Builds an Excel workbook for the post-study ALEI analysis."""

    def __init__(self, *, database: Database) -> None:
        self._database = database

    def build_export(self) -> tuple[str, bytes]:
        """
        Returns (filename, xlsx_bytes).
        Filename pattern: "alei_components_YYYYMMDD_HHMMSS.xlsx"
        """
        workbook = Workbook()

        # ── Sheet 1: Learning Gain ──────────────────────────────────────────
        lg_sheet = workbook.active
        lg_sheet.title = "Learning Gain"
        lg_sheet.append([
            "student_id",
            "study_session_id",
            "condition",
            "lesson_ids",
            "pre_test_score",
            "post_test_score",
            "hake_g",
            "pre_test_completed_at_utc",
            "post_test_completed_at_utc",
        ])

        # ── Sheet 2: Usability & Workload ───────────────────────────────────
        uw_sheet = workbook.create_sheet("Usability & Workload")
        uw_sheet.append([
            "student_id",
            "study_session_id",
            "condition",
            "sus_score",
            "tlx_overall",
            "tlx_mental_demand",
            "tlx_effort",
            "tlx_frustration",
            "tlx_temporal_demand",
            "tlx_performance",
            "tlx_physical_demand",
            "submitted_at_utc",
        ])

        # ── Sheet 3: Prediction Probes ──────────────────────────────────────
        pp_sheet = workbook.create_sheet("Prediction Probes")
        pp_sheet.append([
            "student_id",
            "study_session_id",
            "lesson_id",
            "concept_id",
            "probe_id",
            "attempt_index",
            "actual_value",
            "predicted_value",
            "absolute_error",
            "tolerance",
            "is_within_tolerance",
            "occurred_at_utc",
        ])

        # ── Sheet 4: Behavioral ─────────────────────────────────────────────
        beh_sheet = workbook.create_sheet("Behavioral")
        beh_sheet.append([
            "student_id",
            "study_session_id",
            "lesson_id",
            "code_submit_failures",
            "code_submit_successes",
            "hint_requests",
            "hints_shown",
            "recommendations_shown",
            "recommendation_clicks",
            "recommendation_completes",
            "probes_shown",
            "probes_submitted",
        ])

        # ── Sheet 5: Component Surveys ──────────────────────────────────────
        cs_sheet = workbook.create_sheet("Component Surveys")
        cs_sheet.append([
            "student_id",
            "study_session_id",
            "condition",
            "context_trigger",
            "question_text",
            "likert_value",
            "text_value",
            "submitted_at_utc",
        ])

        with self._database.session() as session:
            # ── Load data ───────────────────────────────────────────────────
            sessions = session.exec(
                select(StudySession).order_by(
                    StudySession.student_id, StudySession.started_at_utc
                )
            ).all()

            surveys = session.exec(
                select(StudySessionSurvey).order_by(
                    StudySessionSurvey.student_id, StudySessionSurvey.submitted_at_utc
                )
            ).all()

            probes = session.exec(
                select(PredictionProbeAttempt).order_by(
                    PredictionProbeAttempt.student_id,
                    PredictionProbeAttempt.occurred_at_utc,
                )
            ).all()

            events = session.exec(select(LearningInteractionEvent)).all()

            # Component survey data
            component_responses = session.exec(
                select(SurveyResponse).order_by(
                    SurveyResponse.condition,
                    SurveyResponse.student_id,
                    SurveyResponse.submitted_at_utc,
                )
            ).all()
            component_item_responses = session.exec(
                select(SurveyItemResponse)
            ).all()
            component_items = session.exec(
                select(SurveyTemplateItem)
            ).all()
            component_templates = session.exec(
                select(SurveyTemplate)
            ).all()

        # ── Build lookup: session_id -> list[LearningInteractionEvent] ──────
        events_by_session: dict[str, list[LearningInteractionEvent]] = defaultdict(list)
        for event in events:
            events_by_session[event.session_id].append(event)

        # ── Sheet 1 rows ────────────────────────────────────────────────────
        for ss in sessions:
            lesson_ids = self._decode_json_list(ss.lesson_ids_json)
            session_events = events_by_session.get(ss.id, [])

            pre_score = self._extract_test_score(session_events, ET.PRE_TEST_COMPLETE)
            post_score = self._extract_test_score(session_events, ET.POST_TEST_COMPLETE)

            hake_g: float | None = None
            if pre_score is not None and post_score is not None and pre_score < 100:
                hake_g = round((post_score - pre_score) / (100 - pre_score), 4)

            pre_ts = (
                ss.pre_test_completed_at_utc.isoformat()
                if ss.pre_test_completed_at_utc is not None
                else None
            )
            post_ts = (
                ss.post_test_completed_at_utc.isoformat()
                if ss.post_test_completed_at_utc is not None
                else None
            )

            lg_sheet.append([
                ss.student_id,
                ss.id,
                ss.condition,
                json.dumps(lesson_ids),
                pre_score,
                post_score,
                hake_g,
                pre_ts,
                post_ts,
            ])

        # ── Sheet 2 rows ────────────────────────────────────────────────────
        for survey in surveys:
            uw_sheet.append([
                survey.student_id,
                survey.study_session_id,
                survey.condition,
                survey.sus_score,
                survey.tlx_overall,
                survey.tlx_mental_demand,
                survey.tlx_effort,
                survey.tlx_frustration,
                survey.tlx_temporal_demand,
                survey.tlx_performance,
                survey.tlx_physical_demand,
                survey.submitted_at_utc.isoformat(),
            ])

        # ── Sheet 3 rows ────────────────────────────────────────────────────
        for probe in probes:
            pp_sheet.append([
                probe.student_id,
                probe.study_session_id,
                probe.lesson_id,
                probe.concept_id,
                probe.probe_id,
                probe.attempt_index,
                probe.actual_value,
                probe.predicted_value,
                probe.absolute_error,
                probe.tolerance,
                probe.is_within_tolerance,
                probe.occurred_at_utc.isoformat(),
            ])

        # ── Sheet 4 rows ────────────────────────────────────────────────────
        # Aggregate per (student_id, study_session_id, lesson_id)
        BehKey = tuple[str, str, str]
        beh: dict[BehKey, dict[str, int]] = defaultdict(
            lambda: {
                ET.CODE_SUBMIT_FAILURE: 0,
                ET.CODE_SUBMIT_SUCCESS: 0,
                ET.HINT_REQUEST: 0,
                ET.HINT_SHOWN: 0,
                ET.RECOMMENDATION_SHOWN: 0,
                ET.RECOMMENDATION_CLICK: 0,
                ET.RECOMMENDATION_COMPLETE: 0,
                ET.PROBE_SHOWN: 0,
                ET.PROBE_SUBMITTED: 0,
            }
        )
        for event in events:
            key: BehKey = (event.student_id, event.session_id, event.lesson_id)
            if event.event_type in beh[key]:
                beh[key][event.event_type] += 1
            else:
                # ensure the key exists even if no tracked event type fires
                _ = beh[key]

        for (student_id, session_id, lesson_id), counts in sorted(beh.items()):
            beh_sheet.append([
                student_id,
                session_id,
                lesson_id,
                counts[ET.CODE_SUBMIT_FAILURE],
                counts[ET.CODE_SUBMIT_SUCCESS],
                counts[ET.HINT_REQUEST],
                counts[ET.HINT_SHOWN],
                counts[ET.RECOMMENDATION_SHOWN],
                counts[ET.RECOMMENDATION_CLICK],
                counts[ET.RECOMMENDATION_COMPLETE],
                counts[ET.PROBE_SHOWN],
                counts[ET.PROBE_SUBMITTED],
            ])

        # ── Sheet 5 rows ────────────────────────────────────────────────────
        # Build lookups from loaded data
        item_text_map = {item.id: item.question_text for item in component_items}
        template_trigger_map = {t.id: t.context_trigger for t in component_templates}
        item_responses_by_response: dict[str, list[SurveyItemResponse]] = defaultdict(list)
        for ir in component_item_responses:
            item_responses_by_response[ir.survey_response_id].append(ir)

        for resp in component_responses:
            context_trigger = template_trigger_map.get(resp.survey_template_id, "unknown")
            item_responses = item_responses_by_response.get(resp.id, [])
            for ir in sorted(item_responses, key=lambda x: item_text_map.get(x.item_id, "")):
                cs_sheet.append([
                    resp.student_id,
                    resp.study_session_id,
                    resp.condition,
                    context_trigger,
                    item_text_map.get(ir.item_id, ir.item_id),
                    ir.likert_value,
                    ir.text_value,
                    resp.submitted_at_utc.isoformat(),
                ])

        content = io.BytesIO()
        workbook.save(content)
        content.seek(0)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return f"alei_components_{timestamp}.xlsx", content.getvalue()

    # ── Helpers ─────────────────────────────────────────────────────────────

    @staticmethod
    def _extract_test_score(
        events: list[LearningInteractionEvent],
        event_type: str,
    ) -> float | None:
        for event in events:
            if event.event_type == event_type:
                try:
                    payload = json.loads(event.payload_json or "{}")
                    if "score" in payload:
                        return float(payload["score"])
                except (json.JSONDecodeError, ValueError, TypeError):
                    continue
        return None

    @staticmethod
    def _decode_json_list(raw: str) -> list[Any]:
        try:
            result = json.loads(raw or "[]")
            return result if isinstance(result, list) else []
        except json.JSONDecodeError:
            return []
